"""
Export service for generating procurement orders in various formats
"""
import csv
import io
import logging
from typing import List, Dict, Any
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting procurement orders in various formats"""
    
    def __init__(self, company_name: str):
        self.company_name = company_name
        self.export_date = datetime.now()
    
    def generate_excel_export(self, products: List[Dict[str, Any]], notes: str = '') -> bytes:
        """Generate Excel (XLSX) export of procurement order"""
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Procurement Order"
            
            # Add header information
            header_row = [
                f"Procurement Order - {self.company_name}",
                "",
                "",
                f"Generated: {self.export_date.strftime('%Y-%m-%d %H:%M:%S')}"
            ]
            ws.append(header_row)
            
            # Merge cells for header
            ws.merge_cells('A1:D1')
            
            # Style the header
            header_font = Font(bold=True, size=14)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Add empty row
            ws.append([])
            
            # Add notes if provided
            if notes:
                ws.append(["Order Notes:", notes])
                ws.append([])
            
            # Add column headers
            headers = ["SKU", "Product Name", "Current Stock", "Recommended Quantity", "Notes"]
            ws.append(headers)
            
            # Style column headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            # Add product data
            for product in products:
                row = [
                    product.get('sku', ''),
                    product.get('name', ''),
                    product.get('current_stock', 0),
                    product.get('recommended_quantity', 0),
                    product.get('notes', '')
                ]
                ws.append(row)
                
                # Apply border to data rows
                for col in range(1, len(row) + 1):
                    cell = ws.cell(row=ws.max_row, column=col)
                    cell.border = thin_border
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A4'
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Failed to generate Excel export: {str(e)}")
            raise
    
    def generate_csv_export(self, products: List[Dict[str, Any]], notes: str = '') -> bytes:
        """Generate CSV export of procurement order"""
        try:
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Add header information
            writer.writerow([f"Procurement Order - {self.company_name}"])
            writer.writerow([f"Generated: {self.export_date.strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])  # Empty row
            
            # Add notes if provided
            if notes:
                writer.writerow(["Order Notes:", notes])
                writer.writerow([])  # Empty row
            
            # Add column headers
            writer.writerow(["SKU", "Product Name", "Current Stock", "Recommended Quantity", "Notes"])
            
            # Add product data
            for product in products:
                writer.writerow([
                    product.get('sku', ''),
                    product.get('name', ''),
                    product.get('current_stock', 0),
                    product.get('recommended_quantity', 0),
                    product.get('notes', '')
                ])
            
            # Convert to bytes with UTF-8 encoding and BOM for Excel compatibility
            csv_content = output.getvalue()
            return f"\ufeff{csv_content}".encode('utf-8')
            
        except Exception as e:
            logger.error(f"Failed to generate CSV export: {str(e)}")
            raise
    
    def generate_pdf_export(self, products: List[Dict[str, Any]], notes: str = '') -> bytes:
        """Generate PDF export of procurement order"""
        try:
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            elements = []
            
            # Get styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Add title
            title = Paragraph(f"Procurement Order - {self.company_name}", title_style)
            elements.append(title)
            
            # Add generation date
            date_text = Paragraph(f"Generated: {self.export_date.strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
            elements.append(date_text)
            elements.append(Spacer(1, 20))
            
            # Add notes if provided
            if notes:
                notes_title = Paragraph("Order Notes:", styles['Heading2'])
                elements.append(notes_title)
                notes_text = Paragraph(notes, styles['Normal'])
                elements.append(notes_text)
                elements.append(Spacer(1, 20))
            
            # Create data for table
            data = [["SKU", "Product Name", "Current Stock", "Recommended Quantity", "Notes"]]
            
            for product in products:
                data.append([
                    product.get('sku', ''),
                    product.get('name', ''),
                    str(product.get('current_stock', 0)),
                    str(product.get('recommended_quantity', 0)),
                    product.get('notes', '')
                ])
            
            # Create table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Failed to generate PDF export: {str(e)}")
            raise